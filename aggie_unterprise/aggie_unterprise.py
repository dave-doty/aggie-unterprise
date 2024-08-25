from __future__ import annotations
from dataclasses import dataclass
from typing import Iterable, List, Dict, Optional, Union, cast
from pathlib import Path
from openpyxl import load_workbook
from tabulate import tabulate
import datetime
import calendar


def format_currency(amount: float) -> str:
    # The "proper" way to do this is with the locale module,
    # but you need the system to have certain locales installed,
    # and I don't want users to run into those stupid errors,
    # so we just manually format the currency.
    return f"${amount:,.2f}" if amount >= 0 else f"-${abs(amount):,.2f}"


# TODO: don't hardcode header rows; search for them instead

summary_header_row_idx = 17
detail_header_row_idx = 17
date_cell = 'A3'

MARKDOWN_TABLE_FORMATS = ['github', 'pipe']


def remove_suffix_starting_with(string: str, substrings: Iterable[str]) -> str:
    for substring in substrings:
        if substring in string:
            idx = string.index(substring)
            string = string[:idx].strip()
            return string
    return string


def remove_substrings(string: str, substrings: Iterable[str]) -> str:
    # make a defensive copy since we modify the list
    substrings = list(substrings)

    # sort substrings by substring relation to each other to avoid removing 
    # a subsubstring that causes the remaining part not to be removed
    # e.g., if we have "abc" and "a", we want to remove "abc" first, otherwise
    # if string is "a abc My Project", we will end up changing like this:
    # " bc My Project" instead of " My Project" because removing the 
    # subsubstring "a" first changes the other substring "abc" to "bc"
    for i, sub1 in enumerate(substrings):
        for j in range(i + 1, len(substrings)):
            sub2 = substrings[j]  # type: ignore #PyCharm thinks list has no [] operator
            if sub1 in sub2:
                substrings[i], substrings[j] = sub2, sub1

    for substring in substrings:
        if substring in string:
            string = string.replace(substring, '')
    return string


def clean_whitespace(string: str) -> str:
    return ' '.join(string.split())


def find_expenses_by_category(summary: ProjectSummary, ws_detail, project_name_header: str, project_name: str) -> None:
    detail_expense_category_header = 'Expenditure Category Name'
    detail_expenses_header = 'Expenses'
    detail_budget_header = 'Budget'
    detail_header_row = ws_detail[detail_header_row_idx]
    detail_col_idxs = {}
    for i, cell in enumerate(detail_header_row):
        header = cell.value
        if header in [project_name_header, detail_expense_category_header, detail_expenses_header,
                      detail_budget_header]:
            detail_col_idxs[header] = i

    found_project_name = False
    for row in ws_detail.iter_rows(min_row=18):
        if row[detail_col_idxs[project_name_header]].value == project_name:
            found_project_name = True
            break
    if not found_project_name:
        raise ValueError(f"Couldn't find expenses for project {project_name}")

    for row in ws_detail.iter_rows(min_row=18):
        if row[detail_col_idxs[project_name_header]].value == project_name:
            category = row[detail_col_idxs[detail_expense_category_header]].value
            if 'Salaries and Wages' in category:
                summary.salary = row[detail_col_idxs[detail_expenses_header]].value
            elif 'Travel' in category:
                summary.travel = row[detail_col_idxs[detail_expenses_header]].value
            elif 'Supplies / Services / Other Expenses' in category:
                summary.supplies = row[detail_col_idxs[detail_expenses_header]].value
            elif 'Fringe Benefits' in category:
                summary.fringe = row[detail_col_idxs[detail_expenses_header]].value
            elif 'Fellowship & Scholarships' in category:
                summary.fellowship = row[detail_col_idxs[detail_expenses_header]].value
            elif 'Indirect Costs' in category:
                summary.indirect = row[detail_col_idxs[detail_expenses_header]].value
            else:
                print(f"Unknown category {category}; consider adding it to the list of categories?")


def extract_date(ws_summary) -> datetime.datetime:
    raw = ws_summary[date_cell].value
    raw = raw.replace('Report Run Date:', '').strip()
    date_and_time = datetime.datetime.strptime(raw, '%Y-%m-%d %I:%M:%S %p')
    return date_and_time


POSSIBLE_HEADERS = ['Expenses', 'Salary', 'Travel', 'Supplies', 'Fringe', 'Fellowship', 'Indirect', 'Balance', 'Budget']


@dataclass
class Summary:
    """
    A summary of all the projects in the AggieEnterprise Excel file created by following [these instructions](
    https://servicehub.ucdavis.edu/servicehub?id=ucd_kb_article&sys_id=cc1942f61b32c6d80e0b2068b04bcbbf).

    Create a Summary object by calling [`Summary.from_file`][aggie_unterprise.Summary.from_file]
    with the filename of the Excel file:

    ```python
    summary = Summary.from_file('2024-8-5.xlsx')
    ```
    """

    project_summaries: List[ProjectSummary]
    """A list of [`ProjectSummary`][aggie_unterprise.ProjectSummary]'s, one for each project found in the
    AggieEnterprise Excel file read by [`Summary.from_file`][aggie_unterprise.Summary.from_file]"""

    date_and_time: datetime.datetime
    """The date and time the summary was generated"""

    @staticmethod
    def from_file(
            fn: Union[str, Path],
            substrings_to_clean: Iterable[str] = (),
            suffixes_to_clean: Iterable[str] = (),
    ) -> Summary:
        """
        Read the Excel file named `fn` (alternately `fn` can be a
        [`pathlib.Path`](https://docs.python.org/3/library/pathlib.html) object)
        and return a list of summaries of projects in the file.

        Args:
            fn: The filename (or [`pathlib.Path`](<https://docs.python.org/3/library/pathlib.html>) object)
                of the AggieExpense Excel file to read.

            substrings_to_clean: A list of substrings to remove from the project names

            suffixes_to_clean: A list of substrings to remove from the project names, including
                the whole suffix following the substring

        Returns:
            A [`Summary`][aggie_unterprise.Summary] object containing the summaries of all the projects in the file.
        """
        if isinstance(fn, Path):
            fn = str(fn.resolve())
        wb = load_workbook(filename=fn, read_only=True)
        ws_summary = wb['Summary']
        ws_detail = wb['Detail']

        date_report_created = extract_date(ws_summary)

        header_row = ws_summary[summary_header_row_idx]
        assert header_row[0].value == "Award Number"
        project_name_header = 'Project Name'
        budget_header = 'Budget'
        expenses_header = 'Expenses'
        balance_header = 'Budget Balance (Budget – (Comm & Exp))'
        col_names = [project_name_header, budget_header, expenses_header, balance_header]

        project_type_header = 'Project Type'

        # used to distinguish CS-specific accounts that all have the same project name of
        # "David Doty ENGR COMPUTER SCIENCE PPM Only"
        # In these cases we grab the Task name and used that as the project name instead
        task_header = 'Task/Subtask Name'

        summary_col_idxs = {}
        for i, cell in enumerate(header_row):
            header = cell.value
            if header in col_names + [task_header] + [project_type_header]:
                summary_col_idxs[header] = i

        project_names = []
        project_summaries = []
        clean_name_to_orig: Dict[str, str] = {}
        for row in ws_summary.iter_rows(min_row=summary_header_row_idx + 1):
            project_name: str = cast(str, row[summary_col_idxs[project_name_header]].value)
            budget_str: str = cast(str, row[summary_col_idxs[budget_header]].value)
            expenses_str: str = cast(str, row[summary_col_idxs[expenses_header]].value)
            balance_str: str = cast(str, row[summary_col_idxs[balance_header]].value)
            if None in [project_name, budget_str, expenses_str]:
                # some rows are empty; skip them
                continue

            budget = float(budget_str)
            expenses = float(expenses_str)
            balance = float(balance_str)
            summary = ProjectSummary(project_name=project_name, balance=balance, budget=budget, expenses=expenses)
            find_expenses_by_category(summary, ws_detail, project_name_header, project_name)

            if project_name in project_names:
                raise ValueError(f'There are duplicates in the project names: "{project_name}" appears in two '
                                 f'different rows of the Summary worksheet of the Excel file "{fn}". '
                                 f'I do not know how to process such a file.')

            # Internal project names are typically of the form "David Doty ENGR COMPUTER SCIENCE PPM Only"
            # so we replace those with the more specific task name such as
            # "CS INDIRECT COST RETURN DOTY DEFAULT PROJECT 13U00" or
            # "CS DISCRETIONARY FUNDS DOTY DEFAULT PROJECT 13U00"

            project_type = row[summary_col_idxs[project_type_header]].value
            assert project_type in ['Internal', 'Sponsored']
            internal = project_type == 'Internal'
            if internal:
                if 'PPM Only' not in project_name:
                    print("WARNING: Internal project name typically contain the phrase 'PPM Only'; "
                          f'double-check that project "{project_name}" is internal and is the correct project name.')
                # replace with more specific task name
                project_name = cast(str, row[summary_col_idxs[task_header]].value)

            clean_project_name = project_name
            clean_project_name = remove_suffix_starting_with(clean_project_name, suffixes_to_clean)
            clean_project_name = remove_substrings(clean_project_name, substrings_to_clean)
            clean_project_name = clean_whitespace(clean_project_name)

            if clean_project_name in clean_name_to_orig.keys():
                # check if iterable suffixes_to_clean is empty
                first_orig_project_name = clean_name_to_orig[clean_project_name]
                if any(True for _ in suffixes_to_clean) or any(True for _ in substrings_to_clean):
                    msg = (f'Warning: After: cleaning up the project names, there are duplicates.'
                           f'The original project names "{first_orig_project_name}" and "{project_name}" '
                           f'both map to the cleaned project name {clean_project_name}. '
                           f'Try specifying different values in `substrings_to_clean` or `suffixes_to_clean`.')
                    raise ValueError(msg)
                else:
                    url = r'https://github.com/dave-doty/aggie-unterprise/issues'
                    assert f'Should be unreachable; report this error to the aggie_unterprise issues page: {url}'
            project_names.append(clean_project_name)

            summary.project_name = clean_project_name

            project_summaries.append(summary)
            clean_name_to_orig[clean_project_name] = project_name

        return Summary(project_summaries, date_report_created)

    def __str__(self) -> str:
        return self.table()

    def __repr__(self) -> str:
        return self.table()

    def table(self, tablefmt: str = 'rounded_outline', headers: Optional[Iterable[str]] = None) -> str:
        """
        Return a representation of the summary as a string in tabular form.

        Args:
            tablefmt: The format of the table; see the Python package\
            [tabulate documentation](<https://github.com/astanin/python-tabulate#table-format>) for options.

            headers: A list of headers to include in the table; must be a subset of
                `{'Expenses', 'Salary', 'Travel', 'Supplies', 'Fringe', 'Fellowship', 'Indirect', 'Balance', 'Budget'}`
                The headers will be displayed in the order they are given, so is also a way to reorder them
                from the default order even if you include all of them.
                If not specified, all headers will be included in the default order.

        Returns:
            A representation of the summary as a string in tabular form
        """
        if headers is None:
            headers = POSSIBLE_HEADERS
        for header in headers:
            if header not in POSSIBLE_HEADERS:
                raise ValueError(f"Invalid heading: {header}; must be one of {', '.join(POSSIBLE_HEADERS)}")

        table = []
        for project_summary in self.project_summaries:
            header_to_field = {
                'Expenses': project_summary.expenses,
                'Salary': project_summary.salary,
                'Travel': project_summary.travel,
                'Supplies': project_summary.supplies,
                'Fringe': project_summary.fringe,
                'Fellowship': project_summary.fellowship,
                'Indirect': project_summary.indirect,
                'Balance': project_summary.balance,
                'Budget': project_summary.budget,
            }
            row = [project_summary.project_name]
            for header in headers:
                row.append(format_currency(header_to_field[header]))

            if tablefmt in MARKDOWN_TABLE_FORMATS:
                # escape $ so markdown does not interpret it as Latex
                for i in range(1, len(row)):
                    row[i] = row[i].replace('$', r'\$')  # type:ignore #PyCharm thinks list has no [] operator
            table.append(row)

        new_headers = ['Project Name'] + list(headers)
        colalign = ['left'] + ['right'] * (len(new_headers) - 1)
        table_tabulated = tabulate(table, headers=new_headers, tablefmt=tablefmt, colalign=colalign)
        return table_tabulated

    def diff_table(self, summary_earlier: Summary, tablefmt: str = 'rounded_outline',
                   headers: Optional[Iterable[str]] = None) -> str:
        """
        Return a representation of the differences between this summary and `summary_earlier`.

        Any project listed in one summary but not the other will be treated as though the other summary
        had a ProjectSummary object with all fields set to 0. For instance, if a new grant was added to the
        later summary, the row for that project will show all the fields as the amounts in the later summary.
        If a grant was removed from the later summary, the row for that project will show all the fields as
        the negative of the amounts in the earlier summary.

        Args:
            summary_earlier: The earlier [`Summary`][aggie_unterprise.Summary] object to compare against

            tablefmt: The format of the table; see the Python package\
            [tabulate documentation](<https://github.com/astanin/python-tabulate#table-format>) for options.

            headers: A list of headers to include in the diff table; must be a subset of
                `{'Expenses', 'Salary', 'Travel', 'Supplies', 'Fringe', 'Fellowship', 'Indirect', 'Balance'}`
                (Note that 'Budget' is not included since it should always equal between two summaries.)
                The headers will be displayed in the order they are given, so is also a way to reorder them
                from the default order even if you include all of them.
                If not specified, all headers will be included in the default order.

        Returns:
            A representation of the summary of differences as a string in tabular form
        """
        if headers is None:
            headers = POSSIBLE_HEADERS
        headers = list(headers)
        for header in headers:
            if header not in POSSIBLE_HEADERS:
                raise ValueError(f"Invalid heading: {header}; must be one of {', '.join(POSSIBLE_HEADERS)}")

        # some projects may have been added or removed between the two summaries
        # in these cases we will let the diff between a project in one that is missing from the other
        # be as though the other had the ProjectSummary but with all fields set to 0
        names_to_projects_later = {summary.project_name: summary for summary in self.project_summaries}
        names_to_projects_earlier = {summary.project_name: summary for summary in summary_earlier.project_summaries}
        names_only_earlier = [name for name in names_to_projects_earlier if name not in names_to_projects_earlier]
        all_project_names = list(names_to_projects_later.keys()) + names_only_earlier

        table = []
        for project_name in all_project_names:
            assert project_name in names_to_projects_later or project_name in names_to_projects_earlier
            null_project_summary = ProjectSummary(project_name=project_name)
            proj_summary_later = names_to_projects_later.get(project_name, null_project_summary)
            proj_summary_earlier = names_to_projects_earlier.get(project_name, null_project_summary)
            assert proj_summary_later is not null_project_summary or proj_summary_earlier is not null_project_summary

            if proj_summary_later.project_name != proj_summary_earlier.project_name:
                raise ValueError("Can't diff summaries with different project names")
            diff = proj_summary_later.diff(proj_summary_earlier)
            header_to_field = {
                'Expenses': diff.expenses,
                'Salary': diff.salary,
                'Travel': diff.travel,
                'Supplies': diff.supplies,
                'Fringe': diff.fringe,
                'Fellowship': diff.fellowship,
                'Indirect': diff.indirect,
                'Balance': diff.balance,
            }
            row = [diff.project_name]
            for header in headers:
                if header == 'Budget':  # don't show budget diff since it's always equal between two summaries
                    continue
                row.append(format_currency(header_to_field[header]))

            if tablefmt in MARKDOWN_TABLE_FORMATS:
                # escape $ so markdown does not interpret it as Latex
                for i in range(1, len(row)):
                    # PyCharm thinks list has no [] operator
                    row[i] = row[i].replace('$', r'\$')  # type:ignore
            table.append(row)

        new_headers = ['Project Name'] + headers
        num_expense_cols = len(headers) - 1 if 'Budget' in headers else len(headers)
        colalign = ['left'] + ['right'] * num_expense_cols
        table_tabulated = tabulate(table, headers=new_headers, tablefmt=tablefmt, colalign=colalign)
        # TODO: when tabulate updates to verion 0.9.1, uncomment the following line
        # ,colglobalalign='left')
        return table_tabulated

    def year(self) -> int:
        """The year of the summary (as an integer)"""
        return self.date_and_time.year

    def month(self) -> str:
        """The month of the summary (as a string)"""
        return calendar.month_name[self.date_and_time.month]

    def day(self) -> int:
        """The day of the summary (as an integer)"""
        return self.date_and_time.day


@dataclass
class ProjectSummary:
    """
    Objecting summarizing a single project, with a name and several dollar amounts gathered from the
    AggieEnterprise Excel file when the ProjectSummary object is constructed when calling
    [`Summary.from_file`][aggie_unterprise.Summary.from_file].
    """

    project_name: str
    """Name of the project. Comes from the column "Project Name" if the project is sponsored 
    (through an external grant) and from the column "Task/Subtask Name" if the project is internal."""

    balance: float = 0
    """The balance of the project. Comes from the column "Budget Balance (Budget – (Comm & Exp))"."""

    budget: float = 0
    """The total budget of the project. Comes from the column "Budget"."""

    expenses: float = 0
    """The total expenses of the project. Comes from the column "Expenses" The other specific types of expenses
    (salary, travel, etc.) should add up to this number."""

    salary: float = 0
    """The salary expenses of the project. Comes from the column "Salaries and Wages" in the Detail worksheet."""

    travel: float = 0
    """The travel expenses of the project. Comes from the column "Travel" in the Detail worksheet."""

    supplies: float = 0
    """The supplies expenses of the project. Comes from the column "Supplies / Services / Other Expenses" 
    in the Detail worksheet."""

    fringe: float = 0
    """The fringe benefits expenses of the project. Comes from the column "Fringe Benefits" in the Detail worksheet."""

    fellowship: float = 0
    """The fellowship and scholarships expenses of the project. Comes from the column "Fellowship & Scholarships"""

    indirect: float = 0
    """The indirect costs of the project. Comes from the column "Indirect Costs" in the Detail worksheet."""

    def diff(self, other: ProjectSummary) -> ProjectSummary:
        """
        Return a new ProjectSummary object that is the difference between this ProjectSummary and `other`.

        Args:
            other: The other ProjectSummary object to compare against

        Returns:
            A new ProjectSummary object that is the difference between this ProjectSummary and `other`.
        """
        if self.project_name != other.project_name:
            raise ValueError("Can't diff summaries with different project names")
        return ProjectSummary(
            project_name=self.project_name,
            balance=self.balance - other.balance,
            budget=0,
            expenses=self.expenses - other.expenses,
            salary=self.salary - other.salary,
            travel=self.travel - other.travel,
            supplies=self.supplies - other.supplies,
            fringe=self.fringe - other.fringe,
            fellowship=self.fellowship - other.fellowship,
            indirect=self.indirect - other.indirect,
        )
